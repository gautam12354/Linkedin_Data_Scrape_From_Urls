from openpyxl import load_workbook
wb = load_workbook("Excel data for list.xlsx")  # Work Book
ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
column = ws['A']  # Column
column_list = [column[x].value for x in range(len(column))]
print(column_list)